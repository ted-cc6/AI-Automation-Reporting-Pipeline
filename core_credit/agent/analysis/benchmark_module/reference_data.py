"""Loaders for External Benchmarks.xlsx.

The "60 DB Benchmarks" tab has a two-row header where the top row carries
merged group labels (e.g. "2024 MFI Index" spanning three country columns)
and the bottom row carries the specific country name for each of those
columns. Rather than depend on openpyxl's merged-cell metadata (which needs
a non-read-only load and is easy to get subtly wrong), we forward-fill the
top header row exactly the way a merged cell reads visually -- the same
technique proved out on the PPI scorecards' forward-filled Dimension column.

Units: every value in both sheets is a 0-1 fraction (a plain share), except
national poverty rates get converted to percentage points (x100) on load,
to match the scale ppi_module.pipeline already reports portfolio poverty
likelihood on. NPS's score-vs-fraction scale question is handled in
lookup.py, not here -- this module stores every value exactly as the sheet
has it.
"""

from __future__ import annotations

import os
import re
import warnings
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Optional

import openpyxl

_YEAR_RE = re.compile(r"(20\d{2})")
_REGION_RE = re.compile(r"Regional Benchmark \(([^)]+)\)")

# Explicit opt-in for a deliberately benchmark-free run. Without it, a missing workbook is a
# hard error (see _require_workbook). Named, not a silent default, on purpose.
_ALLOW_MISSING_ENV = "CORE_CREDIT_ALLOW_MISSING_BENCHMARKS"


def _require_workbook(benchmarks_path: str) -> bool:
    """True when the workbook is present. When it is absent this raises FileNotFoundError --
    UNLESS the operator has explicitly set CORE_CREDIT_ALLOW_MISSING_BENCHMARKS, in which case
    it warns and returns False so the caller degrades every indicator to not_available_reason.

    This used to silently `return ()` on a missing file. That let a whole Core Credit run
    finish and ship a report with EVERY MFI Index benchmark missing, surfaced nowhere until a
    reader noticed -- the same "fail after the fact" failure the Insurance pipeline records as
    R-040 / R-045. The workbook is now committed at `core_credit/External Benchmarks.xlsx`, so
    its absence is a real misconfiguration, not the expected BYOK state the old comment
    described.
    """
    if Path(benchmarks_path).exists():
        return True
    if os.environ.get(_ALLOW_MISSING_ENV):
        warnings.warn(
            f"External Benchmarks.xlsx not found at {benchmarks_path!r} and {_ALLOW_MISSING_ENV} "
            f"is set -- every MFI Index benchmark will be reported as not available for this run.",
            RuntimeWarning,
            stacklevel=3,
        )
        return False
    raise FileNotFoundError(
        f"External Benchmarks.xlsx not found at {benchmarks_path!r}. It ships committed at "
        f"core_credit/External Benchmarks.xlsx and every Core Credit run needs it for the MFI "
        f"Index comparison. Fix the caller's benchmarks_path if this location is wrong; to run "
        f"deliberately without benchmarks, set {_ALLOW_MISSING_ENV}=1."
    )


def _forward_fill(values: list) -> list:
    """Mimics reading a merged cell: carries the last non-null value forward across nulls."""
    filled = []
    current = None
    for v in values:
        if v is not None:
            current = v
        filled.append(current)
    return filled


def _clean_text(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


@dataclass(frozen=True)
class MfiIndexColumn:
    index: int
    kind: str  # "global" | "regional" | "country"
    key: Optional[str]  # region name (sheet's own spelling) or country name; None for global
    year: Optional[int]


def _parse_mfi_index_columns(header_row0: tuple, header_row1: tuple) -> list:
    filled_group_labels = _forward_fill(list(header_row0))
    columns = []
    for i, (raw_label, group_label, country_name) in enumerate(zip(header_row0, filled_group_labels, header_row1)):
        if country_name:
            year_match = _YEAR_RE.search(group_label or "")
            columns.append(
                MfiIndexColumn(
                    index=i, kind="country", key=str(country_name).strip(), year=int(year_match.group(1)) if year_match else None
                )
            )
            continue
        if raw_label and str(raw_label).startswith("Regional Benchmark"):
            region_match = _REGION_RE.search(str(raw_label))
            year_match = _YEAR_RE.search(str(raw_label))
            columns.append(
                MfiIndexColumn(
                    index=i,
                    kind="regional",
                    key=region_match.group(1).strip() if region_match else None,
                    year=int(year_match.group(1)) if year_match else None,
                )
            )
            continue
        if raw_label and str(raw_label).startswith("Global Benchmark"):
            year_match = _YEAR_RE.search(str(raw_label))
            columns.append(MfiIndexColumn(index=i, kind="global", key=None, year=int(year_match.group(1)) if year_match else None))
    return columns


@dataclass(frozen=True)
class MfiIndexIndicatorRow:
    dimension: Optional[str]
    indicator_name: str
    internal_definition: Optional[str]
    external_definition: Optional[str]
    comment: Optional[str]
    global_value: Optional[float]
    global_year: Optional[int]
    regional_values: dict = field(default_factory=dict)  # sheet region name -> (value, year)
    country_values: dict = field(default_factory=dict)  # country name -> (value, year)


@lru_cache(maxsize=8)
def load_mfi_index_sheet(benchmarks_path: str) -> tuple:
    """Every indicator row from the "60 DB Benchmarks" tab, values keyed by the level they were reported at.

    Raises FileNotFoundError if `benchmarks_path` doesn't exist, so a misconfigured path fails
    the run early rather than silently shipping a report with every MFI Index benchmark
    missing. Set CORE_CREDIT_ALLOW_MISSING_BENCHMARKS to opt into a benchmark-free run, in
    which case this returns an empty tuple and every indicator degrades to
    not_available_reason -- see _require_workbook.
    """
    if not _require_workbook(benchmarks_path):
        return ()
    wb = openpyxl.load_workbook(Path(benchmarks_path), data_only=True, read_only=True)
    ws = wb["60 DB Benchmarks"]
    rows = list(ws.iter_rows(values_only=True))
    header_row0, header_row1 = rows[0], rows[1]
    columns = _parse_mfi_index_columns(header_row0, header_row1)
    data_rows = rows[2:]
    dimensions = _forward_fill([r[0] for r in data_rows])

    out = []
    for i, r in enumerate(data_rows):
        indicator_name = _clean_text(r[1]) if len(r) > 1 else None
        if not indicator_name:
            continue
        global_value, global_year = None, None
        regional_values: dict = {}
        country_values: dict = {}
        for col in columns:
            value = r[col.index] if col.index < len(r) else None
            if value is None:
                continue
            value = float(value)
            if col.kind == "global":
                global_value, global_year = value, col.year
            elif col.kind == "regional" and col.key:
                regional_values[col.key] = (value, col.year)
            elif col.kind == "country" and col.key:
                country_values[col.key] = (value, col.year)

        out.append(
            MfiIndexIndicatorRow(
                dimension=dimensions[i],
                indicator_name=indicator_name,
                internal_definition=_clean_text(r[2]) if len(r) > 2 else None,
                external_definition=_clean_text(r[3]) if len(r) > 3 else None,
                comment=_clean_text(r[16]) if len(r) > 16 else None,
                global_value=global_value,
                global_year=global_year,
                regional_values=regional_values,
                country_values=country_values,
            )
        )
    return tuple(out)


@dataclass(frozen=True)
class NationalPovertyRateRow:
    region: Optional[str]
    country_name: str
    source: Optional[str]
    year: Optional[int]
    rates: dict = field(default_factory=dict)  # line_code -> Optional[float], percentage points (0-100)


@lru_cache(maxsize=8)
def load_national_poverty_rates(benchmarks_path: str) -> tuple:
    """Every country row from the "PPI Benchmarks" tab, converted from 0-1 fractions to percentage
    points so it's directly comparable with ppi_module's CountryPovertyResult.values.

    Raises FileNotFoundError if `benchmarks_path` doesn't exist, unless
    CORE_CREDIT_ALLOW_MISSING_BENCHMARKS is set -- see load_mfi_index_sheet's docstring and
    _require_workbook.
    """
    if not _require_workbook(benchmarks_path):
        return ()
    wb = openpyxl.load_workbook(Path(benchmarks_path), data_only=True, read_only=True)
    ws = wb["PPI Benchmarks"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    line_codes = list(header[4:])
    data_rows = rows[1:]
    regions = _forward_fill([r[0] for r in data_rows])

    out = []
    for i, r in enumerate(data_rows):
        country_name = _clean_text(r[1]) if len(r) > 1 else None
        if not country_name:
            continue
        rates = {}
        for j, code in enumerate(line_codes, start=4):
            if code is None or j >= len(r):
                continue
            value = r[j]
            rates[code] = (float(value) * 100) if value is not None else None
        out.append(
            NationalPovertyRateRow(
                region=regions[i],
                country_name=country_name,
                source=_clean_text(r[2]) if len(r) > 2 else None,
                year=int(r[3]) if len(r) > 3 and r[3] is not None else None,
                rates=rates,
            )
        )
    return tuple(out)
