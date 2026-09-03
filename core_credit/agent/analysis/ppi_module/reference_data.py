"""Loaders for the PPI_scorecards.xlsx / PPI_lookups.xlsx reference workbooks.

Key design decision (see the project discussion this came out of): we never
match a client's answer against the scorecard's `Answer_Option` ID string
(e.g. `ppi01-RWA-v1-R-0010`) -- that column has inconsistent zero-padding
between countries and question ranges (confirmed on Rwanda's district
question, options 10+ jump from 3-digit to 4-digit padding) and was the
root cause of a real scoring bug VisionFund hit before. Instead we parse the
letter/label prefix directly off the scorecard's own `OptionValues` text
(e.g. "aa) Kayonza" -> "AA", "J. Imbabura" -> "J") and match that against the
survey's own `resp_value` field, which is already that same prefix,
uppercased. This is robust regardless of which padding or numbering
convention a given country's guide happens to use.
"""

from __future__ import annotations

import os
import re
import warnings
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Optional

import openpyxl

_PREFIX_RE = re.compile(r"^\s*([A-Za-z]{1,3})[.)]")

# Explicit opt-in for a run without the PPI reference workbooks. Without it a missing workbook
# is a hard error (see _require_workbook) -- mirrors benchmark_module's
# CORE_CREDIT_ALLOW_MISSING_BENCHMARKS. Setting this makes every country score NOT_AVAILABLE and
# Part 2 report no poverty likelihood at all, so it is a deliberate, loud choice, not a default.
_ALLOW_MISSING_ENV = "CORE_CREDIT_ALLOW_MISSING_PPI"


def _require_workbook(path: str) -> bool:
    """True when the workbook is present. When it is absent this raises FileNotFoundError --
    UNLESS CORE_CREDIT_ALLOW_MISSING_PPI is set, in which case it warns and returns False so the
    caller yields an empty scorecard/lookup (every country -> NOT_AVAILABLE downstream).

    PPI_scorecards.xlsx / PPI_lookups.xlsx used to resolve to a path that did not exist after the
    project folder was renamed (CC-009 / the same defect for External Benchmarks.xlsx). openpyxl
    would then raise a bare FileNotFoundError deep in a country loop, or -- worse for a reviewer
    trying to trust a low n_scored figure -- leave the real reason for an exclusion ambiguous.
    Both workbooks now ship committed at core_credit/.
    """
    if Path(path).exists():
        return True
    if os.environ.get(_ALLOW_MISSING_ENV):
        warnings.warn(
            f"PPI reference workbook not found at {path!r} and {_ALLOW_MISSING_ENV} is set -- "
            f"every country's poverty likelihood will be reported as not available for this run.",
            RuntimeWarning,
            stacklevel=3,
        )
        return False
    raise FileNotFoundError(
        f"PPI reference workbook not found at {path!r}. PPI_scorecards.xlsx and PPI_lookups.xlsx "
        f"ship committed at core_credit/ and Part 2 (Poverty Likelihood) needs both. Fix the "
        f"caller's path if this location is wrong; to run deliberately without PPI scoring, set "
        f"{_ALLOW_MISSING_ENV}=1."
    )

# Every poverty-line / benchmark-line column that appears across the country
# tabs. Country guides only populate a subset of these.
LINE_CODES = [
    "National100",
    "USD125day2005PPP",
    "USD250day2005PPP",
    "USD500day2005PPP",
    "USD190day2011PPP",
    "USD320day2011PPP",
    "USD550day2011PPP",
    "USD215day2017PPP",
    "USD365day2017PPP",
    "USD685day2017PPP",
    "Bottom20thPercentile",
    "Bottom40thPercentile",
    "Bottom60thPercentile",
    "Bottom80thPercentile",
    "ExtremePovertyLine",
]


def parse_option_prefix(label: Optional[str]) -> Optional[str]:
    """Extracts the leading letter code from a scorecard OptionValues label.

    'aa) Kayonza' -> 'AA', 'J. Imbabura' -> 'J', 'A.Quito' -> 'A', 'Seis o más' -> None.
    """
    if not label:
        return None
    m = _PREFIX_RE.match(str(label))
    return m.group(1).upper() if m else None


@dataclass(frozen=True)
class ScorecardOption:
    guide_id: str
    survey_version: int
    question: int
    option_order: int
    option_prefix: Optional[str]
    option_label: str
    points: dict  # line_code -> Optional[float]


@dataclass(frozen=True)
class LookupRow:
    guide_id: str
    score: int
    likelihoods: dict  # line_code -> Optional[float]
    record_modified_date: Optional[object] = None
    record_created_date: Optional[object] = None


def _open_workbook(path: str):
    """Returns an openpyxl workbook, or None when the file is legitimately absent and
    CORE_CREDIT_ALLOW_MISSING_PPI is set. Raises FileNotFoundError otherwise -- see
    _require_workbook.
    """
    if not _require_workbook(path):
        return None
    return openpyxl.load_workbook(Path(path), data_only=True, read_only=True)


@lru_cache(maxsize=128)
def load_scorecard(country_code: str, scorecard_path: str) -> tuple:
    """Every answer-option row for `country_code`, across every guide version in the tab."""
    wb = _open_workbook(scorecard_path)
    if wb is None or country_code not in wb.sheetnames:
        return tuple()
    ws = wb[country_code]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return tuple()
    header = rows[0]
    idx = {h: i for i, h in enumerate(header) if h is not None}
    # Myanmar and Malawi's tabs name this column "QuestionOptionValues" instead of
    # "OptionValues" like every other country -- an inconsistency in the source
    # workbook itself, not a naming choice we get to make.
    option_values_col = "OptionValues" if "OptionValues" in idx else "QuestionOptionValues"
    required_cols = {"PPI_Guide_Id", "SurveyVersion", "PPI_question", "QuestionOptionOrder"}
    if not required_cols.issubset(idx) or option_values_col not in idx:
        return tuple()
    line_idx = {code: idx[code] for code in LINE_CODES if code in idx}

    options = []
    for r in rows[1:]:
        guide_id = r[idx["PPI_Guide_Id"]]
        survey_version = r[idx["SurveyVersion"]]
        question = r[idx["PPI_question"]]
        option_order = r[idx["QuestionOptionOrder"]]
        if guide_id is None or survey_version is None or question is None or option_order is None:
            continue
        label = r[idx[option_values_col]]
        points = {code: r[i] for code, i in line_idx.items()}
        options.append(
            ScorecardOption(
                guide_id=str(guide_id),
                survey_version=int(survey_version),
                question=int(question),
                option_order=int(option_order),
                option_prefix=parse_option_prefix(label),
                option_label="" if label is None else str(label),
                points=points,
            )
        )
    return tuple(options)


@lru_cache(maxsize=128)
def load_lookup(country_code: str, lookup_path: str) -> tuple:
    """Every (guide, score) -> likelihood row for `country_code`.

    Some countries (confirmed on Ecuador) carry two conflicting likelihood
    tables under the same guide_id -- an older one and a refreshed one --
    with no version column to tell them apart. This loader returns every row
    as-is; build_lookup_index (scoring.py) is what resolves the ambiguity,
    by keeping whichever row was modified most recently.
    """
    wb = _open_workbook(lookup_path)
    if wb is None or country_code not in wb.sheetnames:
        return tuple()
    ws = wb[country_code]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return tuple()
    header = rows[0]
    idx = {h: i for i, h in enumerate(header) if h is not None}
    if "PPI_Guide_Id" not in idx or "PPI_Score" not in idx:
        return tuple()
    line_idx = {code: idx[code] for code in LINE_CODES if code in idx}

    out = []
    for r in rows[1:]:
        guide_id = r[idx["PPI_Guide_Id"]]
        score = r[idx["PPI_Score"]]
        if guide_id is None or score is None:
            continue
        likelihoods = {code: r[i] for code, i in line_idx.items()}
        out.append(
            LookupRow(
                guide_id=str(guide_id),
                score=int(score),
                likelihoods=likelihoods,
                record_modified_date=r[idx["Record_Modified_Date"]] if "Record_Modified_Date" in idx else None,
                record_created_date=r[idx["Record_Created_Date"]] if "Record_Created_Date" in idx else None,
            )
        )
    return tuple(out)


def guide_id_for_survey_version(country_code: str, survey_version: int, scorecard_path: str) -> Optional[str]:
    """The scorecard is joined on the client's own SurveyVersion -- never a hardcoded country->guide
    table, since one country (Zambia) has two survey versions that both map to the same guide.
    """
    for opt in load_scorecard(country_code, scorecard_path):
        if opt.survey_version == survey_version:
            return opt.guide_id
    return None


def available_line_codes(country_code: str, guide_id: str, survey_version: int, scorecard_path: str) -> set:
    """Which poverty-line columns actually have point values for this specific guide+version.

    Scoped to survey_version for the same reason build_scorecard_index is
    (scoring.py): different survey versions sharing one guide_id can
    populate a different subset of poverty-line columns (confirmed on
    Ecuador -- the 2017-PPP lines are only populated under survey version 3).
    """
    codes = set()
    for opt in load_scorecard(country_code, scorecard_path):
        if opt.guide_id != guide_id or opt.survey_version != survey_version:
            continue
        for code, value in opt.points.items():
            if value is not None:
                codes.add(code)
    return codes
